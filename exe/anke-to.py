import openpyxl
import os
import sys

fileno = 2 #集計結果は2行目から

testpass = input("集計対象のファイルがあるパスを入力してください。:")
print(testpass)

testhanni = input("集計対象のセルを入力（B54:B59）:")
print(testhanni)

#パスの中にあるファイル名を取得
files = os.listdir(testpass+'\\')
print(files)
print('ファイル数:'+str(len(files)))

hantei = input("集計開始する場合は、1を入力:")
if hantei != '1':
    print('実行を中止します')
    sys.exit(1)

wb_result = openpyxl.Workbook() #集計結果Book ブックハンドラ取得
sheet_result = wb_result.active #集計結果Book シートハンドラ取得
sheet_result.title = '集計結果'
sheet_result.cell(1,1,"ファイル名") #1行目に項目を記載
sheet_result.cell(1,2,testhanni)

for fcnt in files: #取得したファイル数分でループ
    wb = openpyxl.load_workbook(testpass+'\\'+fcnt, data_only=True) #ファイルオープン
    sheet_result.cell(fileno,1,fcnt) #A列にファイル名記載
    ws = wb.active #開いたシートに結果がある前提
    colno = 2  # B列から入力
    for rows in ws[testhanni]:
        sheet_result.cell(fileno, colno, rows[0].value)
        colno+=1
    fileno+=1

wb_result.save('集計結果.xlsx')